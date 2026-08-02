from django.contrib import messages
from django.apps import apps
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.models import Session
from django.core.exceptions import PermissionDenied
from django.db import OperationalError, ProgrammingError, transaction
from django.db.models.deletion import ProtectedError
from django.db.models import CharField, Max, Q, TextField
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from io import BytesIO, StringIO
from urllib.parse import urlencode
import csv
import re
import unicodedata
import bleach

from apps.accounts.models import Empresa, Setor
from .form_registry import (
    FORMULARIOS_CONFIGURAVEIS,
    consultar_campos_formularios,
    opcoes_formularios,
)
from .permissions import role_required

from .forms import EmpresaForm, ModuleForm, ScreenDefinitionForm, ScreenFieldForm
from .locks import adquirir_trava_edicao, liberar_trava_edicao
from .models import (
    Cep,
    ConfiguracaoCampoFormulario,
    IconeSistema,
    Module,
    ScreenDefinition,
    ScreenField,
    TabelaAuxiliarGlobal,
    TipoPrestadorConselho,
    TravaEdicao,
    ValorAuxiliarGlobal,
)
from .table_utils import paginate_table


@login_required
def home(request):
    return render(request, "core/home.html")


def health(request):
    try:
        Empresa.objects.only("pk").first()
    except (OperationalError, ProgrammingError):
        return JsonResponse({"status": "unhealthy", "database": "error"}, status=503)
    return JsonResponse({"status": "ok", "database": "ok"})


def _query_text(request):
    return request.GET.get("q", "").strip().replace("%", "")


def _auxiliary_code(value):
    normalized = unicodedata.normalize("NFD", value)
    normalized = "".join(character for character in normalized if unicodedata.category(character) != "Mn")
    return re.sub(r"[^A-Z0-9]+", "_", normalized.upper()).strip("_")[:40]


@login_required
def placeholder(request):
    return render(request, "core/placeholder.html")


def _empresa_atual(request):
    return get_object_or_404(Empresa, pk=request.session.get("cd_empresa") or 1)


def _dados_trava_requisicao(request):
    tipo = (request.POST.get("tipo") or request.POST.get("table") or "").strip()[:80]
    recurso_id = (request.POST.get("recurso_id") or request.POST.get("id") or "").strip()[:120]
    titulo = (request.POST.get("titulo") or "").strip()[:180]
    if not tipo or not recurso_id:
        return "", "", ""
    return tipo, recurso_id, titulo or f"{tipo} {recurso_id}"


@login_required
def adquirir_trava_generica(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "MÃ©todo nÃ£o permitido."}, status=405)
    tipo, recurso_id, titulo = _dados_trava_requisicao(request)
    if not tipo or not recurso_id:
        return JsonResponse({"ok": False, "error": "Recurso de trava invÃ¡lido."}, status=400)
    resultado = adquirir_trava_edicao(
        _empresa_atual(request),
        request.user,
        tipo,
        recurso_id,
        titulo,
        request.session.session_key or "",
    )
    if not resultado.permitido:
        return JsonResponse({"ok": False, "error": resultado.mensagem}, status=409)
    return JsonResponse({"ok": True})


@login_required
def liberar_trava_generica(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "MÃ©todo nÃ£o permitido."}, status=405)
    tipo, recurso_id, _titulo = _dados_trava_requisicao(request)
    if not tipo or not recurso_id:
        return HttpResponse(status=204)
    liberar_trava_edicao(
        _empresa_atual(request),
        request.user,
        tipo,
        recurso_id,
        motivo="Liberada ao sair da tela.",
    )
    return HttpResponse(status=204)


@login_required
@role_required("TI")
def sessoes_travas(request):
    request.current_tab_title = "TI > Sessões e travas"
    request.current_tab_root_title = "Sessões e travas"
    request.current_module_title = "TI"
    request.current_can_query = False
    request.current_can_save = False
    request.current_can_remove = False
    request.current_reload_url = request.get_full_path()
    empresa = get_object_or_404(Empresa, cd_empresa=request.session.get("cd_empresa") or 1, sn_ativo=True)
    if request.method == "POST":
        trava = get_object_or_404(TravaEdicao, cd_empresa=empresa, pk=request.POST.get("trava"))
        trava.sn_ativa = False
        trava.ds_liberacao = f"Liberada manualmente por {request.user.get_username()}."
        trava.save(update_fields=["sn_ativa", "ds_liberacao", "updated_at"])
        messages.success(request, "Trava liberada com sucesso.")
        return redirect(f"{request.path}?aba=travas")
    agora = timezone.now()
    aba = request.GET.get("aba") or "sessoes"
    if aba not in {"sessoes", "travas"}:
        aba = "sessoes"
    travas = (
        TravaEdicao.objects.select_related("cd_usuario")
        .filter(cd_empresa=empresa, sn_ativa=True)
        .order_by("-updated_at")
    )
    for trava in travas:
        trava.tempo_trava = agora - trava.created_at
        trava.expirada = trava.dh_expiracao < agora

    User = get_user_model()
    sessoes_decodificadas = []
    user_ids = set()
    for sessao in Session.objects.filter(expire_date__gte=agora).order_by("-expire_date"):
        dados = sessao.get_decoded()
        usuario_id = dados.get("_auth_user_id")
        if not usuario_id:
            continue
        cd_empresa_sessao = dados.get("cd_empresa")
        if cd_empresa_sessao and str(cd_empresa_sessao) != str(empresa.pk):
            continue
        user_ids.add(usuario_id)
        sessoes_decodificadas.append((sessao, dados, usuario_id))
    usuarios = User.objects.in_bulk(user_ids)
    travas_por_usuario = {}
    for trava in travas:
        travas_por_usuario.setdefault(str(trava.cd_usuario_id), []).append(trava)

    def data_sessao(valor):
        if not valor:
            return None
        parsed = parse_datetime(str(valor))
        if parsed and timezone.is_naive(parsed):
            parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
        return parsed

    def formatar_duracao(delta):
        total_segundos = max(0, int(delta.total_seconds()))
        horas, resto = divmod(total_segundos, 3600)
        minutos = resto // 60
        if horas:
            return f"{horas}h {minutos:02d}min"
        return f"{minutos}min"

    sessoes = []
    for sessao, dados, usuario_id in sessoes_decodificadas:
        usuario = usuarios.get(usuario_id)
        if not usuario and str(usuario_id).isdigit():
            usuario = usuarios.get(int(usuario_id))
        inicio = data_sessao(dados.get("inicio_sessao_em"))
        ultimo_acesso = data_sessao(dados.get("ultimo_acesso_em"))
        travas_usuario = travas_por_usuario.get(str(usuario_id), [])
        sessoes.append(
            {
                "chave": sessao.session_key,
                "usuario": usuario,
                "usuario_nome": usuario.get_username() if usuario else f"Usuário {usuario_id}",
                "empresa": dados.get("empresa_nome") or empresa.nm_empresa,
                "sistema": dados.get("ultimo_sistema") or "Não rastreado",
                "ultima_tela": dados.get("ultima_tela") or "Não rastreada",
                "ultima_rota": dados.get("ultima_rota") or "-",
                "inicio": inicio,
                "tempo_acesso": formatar_duracao(agora - inicio) if inicio else "-",
                "ultimo_acesso": ultimo_acesso,
                "expira_em": sessao.expire_date,
                "sessao_atual": sessao.session_key == request.session.session_key,
                "travas_ativas": travas_usuario,
                "total_travas": len(travas_usuario),
            }
        )
    return render(
        request,
        "core/sessoes_travas.html",
        {"aba": aba, "sessoes": sessoes, "travas": travas, "agora": agora},
    )


@login_required
@role_required("TI")
def configurar_formularios(request):
    request.current_tab_title = "Global > Formulários > Configurar formulários"
    request.current_tab_root_title = "Configurar formulários"
    request.current_module_title = "Global"
    request.current_can_query = False
    request.current_can_remove = False
    empresa = get_object_or_404(
        Empresa,
        cd_empresa=request.session.get("cd_empresa"),
        sn_ativo=True,
    )
    codigo_formulario = (
        request.POST.get("formulario")
        if request.method == "POST"
        else request.GET.get("formulario", "")
    )
    nome_campo = request.POST.get("nome_campo", "") if request.method == "POST" else request.GET.get("nome_campo", "")
    if codigo_formulario not in FORMULARIOS_CONFIGURAVEIS:
        codigo_formulario = ""
    consultando = request.method == "POST" or request.GET.get("consultar") == "1"
    campos = consultar_campos_formularios(empresa, codigo_formulario, nome_campo) if consultando else []
    request.current_can_query = True
    request.current_start_query = not consultando
    request.current_can_save = bool(campos)
    if request.method == "POST":
        obrigatorios = set(request.POST.getlist("campos_obrigatorios"))
        chaves_resultado = set(request.POST.getlist("campos_resultado"))
        campos_disponiveis = {
            campo["chave"]: campo
            for campo in consultar_campos_formularios(empresa, codigo_formulario, nome_campo)
            if campo["editavel"]
        }
        with transaction.atomic():
            for chave in chaves_resultado | obrigatorios:
                campo = campos_disponiveis.get(chave)
                if not campo:
                    continue
                configuracao, criada = ConfiguracaoCampoFormulario.objects.get_or_create(
                    cd_empresa=empresa,
                    cd_formulario=campo["formulario"],
                    cd_campo=campo["codigo"],
                    defaults={
                        "sn_obrigatorio": chave in obrigatorios,
                        "cd_usuario_criacao": request.user,
                        "cd_usuario_atualizacao": request.user,
                    },
                )
                if not criada:
                    configuracao.sn_obrigatorio = chave in obrigatorios
                    configuracao.cd_usuario_atualizacao = request.user
                    configuracao.save(
                        update_fields=["sn_obrigatorio", "cd_usuario_atualizacao", "updated_at"]
                    )
        messages.success(request, "Configuração do formulário salva e aplicada.")
        parametros = {"consultar": "1"}
        if codigo_formulario:
            parametros["formulario"] = codigo_formulario
        if nome_campo:
            parametros["nome_campo"] = nome_campo
        return redirect(f"{request.path}?{urlencode(parametros)}")
    return render(
        request,
        "core/configurar_formularios.html",
        {
            "formularios": opcoes_formularios(),
            "formulario_selecionado": codigo_formulario,
            "nome_campo": nome_campo,
            "nome_formulario": FORMULARIOS_CONFIGURAVEIS.get(codigo_formulario, {}).get("nome", "Campos encontrados"),
            "campos": campos,
        },
    )


@login_required
def dynamic_screen(request, slug):
    if slug == "ti-alteracao-senha-usuario":
        return redirect("ti:alteracao_senha_usuario")
    screen = get_object_or_404(
        ScreenDefinition.objects.select_related("module").prefetch_related("fields"),
        slug=slug,
        active=True,
        module__active=True,
    )
    screen_path = " > ".join(
        part for part in (screen.module.title, screen.parent_label, screen.title) if part
    )
    request.current_tab_title = screen_path or screen.title
    request.current_tab_root_title = screen.title
    request.current_module_title = screen.module.title
    request.current_can_query = screen.allow_query
    request.current_can_remove = screen.allow_delete
    auxiliary_screens = {
        "cadastros-planos": ("plano", "Planos", "Cadastros"),
        "cadastros-procedimentos": ("procedimento", "Procedimentos", "Cadastros"),
    }
    if slug in auxiliary_screens:
        table_name, title, module_title = auxiliary_screens[slug]
        return global_auxiliary_values(
            request,
            table_name,
            custom_title=title,
            custom_module=module_title,
        )

    template_by_type = {
        ScreenDefinition.TYPE_FORM: "core/dynamic_form.html",
        ScreenDefinition.TYPE_REPORT: "core/dynamic_report.html",
        ScreenDefinition.TYPE_DASHBOARD: "core/dynamic_dashboard.html",
    }
    template = template_by_type.get(screen.screen_type, "core/dynamic_form.html")
    return render(request, template, {"screen": screen, "fields": screen.fields.filter(visible=True)})


def _model_for_table(table_name):
    normalized = (table_name or "").lower()
    if not normalized:
        return None
    for model in apps.get_models():
        meta = model._meta
        if meta.db_table.lower() == normalized or meta.model_name.lower() == normalized:
            return model
    return None


@login_required
def lookup_options(request):
    table_name = request.GET.get("table", "")
    query = request.GET.get("q", "")
    value_field = request.GET.get("value", "")
    display_field = request.GET.get("display", "")
    model = _model_for_table(table_name)
    if not model:
        return JsonResponse({"results": []})

    model_fields = {field.name: field for field in model._meta.fields}
    pk_field = model._meta.pk.name
    value_field = value_field if value_field in model_fields else pk_field
    display_field = display_field if display_field in model_fields else ""
    if not display_field:
        display_field = next(
            (
                field.name
                for field in model._meta.fields
                if isinstance(field, (CharField, TextField)) and field.name != value_field
            ),
            value_field,
        )

    records = model.objects.all()
    if "cd_empresa" in model_fields and model_fields["cd_empresa"].remote_field:
        records = records.filter(cd_empresa_id=request.session.get("cd_empresa") or 1)
    if "sn_ativo" in model_fields:
        records = records.filter(sn_ativo=True)
    if query:
        search_filter = Q()
        for field in model._meta.fields:
            if isinstance(field, (CharField, TextField)):
                search_filter |= Q(**{f"{field.name}__icontains": query.replace("%", "")})
        if search_filter:
            records = records.filter(search_filter)

    results = [
        {
            "value": str(getattr(record, value_field, "")),
            "label": str(getattr(record, display_field, "")),
        }
        for record in records.order_by(display_field)[:20]
    ]
    return JsonResponse({"results": results})


@login_required
@role_required("TI")
def system_screens(request):
    request.current_tab_title = "Global > Configuração do Sistema > Módulos e Telas"
    request.current_tab_root_title = "Módulos e Telas"
    request.current_module_title = "Global"
    request.current_can_query = True
    request.current_can_remove = False
    request.current_new_url = f"{reverse('core:system_screens')}?novo=1"

    if request.GET.get("consultar") == "1":
        modules = Module.objects.all()
        code = request.GET.get("code", "").strip().replace("%", "")
        title = request.GET.get("title", "").strip().replace("%", "")
        icon = request.GET.get("icon", "").strip()
        order = request.GET.get("order", "").strip()
        active = request.GET.get("active", "").strip().lower()
        if code:
            modules = modules.filter(code__icontains=code)
        if title:
            modules = modules.filter(title__icontains=title)
        if icon:
            modules = modules.filter(icon=icon)
        if order.isdigit():
            modules = modules.filter(order=int(order))
        if active in {"true", "false"}:
            modules = modules.filter(active=active == "true")
        result_ids = list(modules.order_by("code", "pk").values_list("pk", flat=True)[:500])
        request.session["consulta_modulos_sistema"] = result_ids
        if not result_ids:
            messages.warning(request, "Nenhum módulo encontrado para os filtros informados.")
            return redirect(f"{reverse('core:system_screens')}?sem_resultados=1")
        return redirect(
            f"{reverse('core:system_screens')}?module={result_ids[0]}&origem=consulta"
        )

    module_id = request.POST.get("module_id") or request.GET.get("module")
    module = get_object_or_404(Module, pk=module_id) if str(module_id or "").isdigit() else None
    query_context = request.GET.get("origem") == "consulta"
    result_ids = request.session.get("consulta_modulos_sistema", []) if query_context else []
    query_mode = bool(request.GET.get("sem_resultados") == "1" or (not module and request.GET.get("novo") != "1"))
    form = ModuleForm(request.POST or None, instance=module, query_mode=query_mode)

    if request.method == "POST" and module and module.is_system:
        raise PermissionDenied("Módulos estruturais só podem ser alterados por migrações do sistema.")
    if request.method == "POST" and form.is_valid():
        saved_module = form.save()
        messages.success(request, "Módulo salvo com sucesso.")
        redirect_params = {"module": saved_module.pk}
        if query_context and saved_module.pk in result_ids:
            redirect_params["origem"] = "consulta"
        return redirect(
            f"{reverse('core:system_screens')}?{urlencode(redirect_params)}"
        )

    if module:
        if module.is_system:
            request.current_can_save = False
        else:
            request.current_toggle_active_url = reverse("core:system_module_toggle_active", args=[module.pk])
            request.current_toggle_active_label = "Desativar" if module.active else "Ativar"
    request.current_start_query = query_mode

    if query_context and module and module.pk in result_ids:
        current_index = result_ids.index(module.pk)
        request.current_record_status = f"Item {current_index + 1} de {len(result_ids)}"
        if current_index > 0:
            request.current_first_url = f"{reverse('core:system_screens')}?module={result_ids[0]}&origem=consulta"
            request.current_previous_url = f"{reverse('core:system_screens')}?module={result_ids[current_index - 1]}&origem=consulta"
        if current_index < len(result_ids) - 1:
            request.current_next_url = f"{reverse('core:system_screens')}?module={result_ids[current_index + 1]}&origem=consulta"
            request.current_last_url = f"{reverse('core:system_screens')}?module={result_ids[-1]}&origem=consulta"

    root_items = (
        module.screens.filter(parent__isnull=True)
        .prefetch_related("children__children__children")
        .order_by("order", "title")
        if module
        else ScreenDefinition.objects.none()
    )
    return render(
        request,
        "core/system_screens.html",
        {"form": form, "module": module, "root_items": root_items},
    )


@login_required
@role_required("TI")
def system_module_toggle_active(request, pk):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método não permitido."}, status=405)
    module = get_object_or_404(Module, pk=pk)
    if module.is_system:
        raise PermissionDenied("Módulos estruturais não podem ser desativados.")
    module.active = not module.active
    module.save(update_fields=["active", "updated_at"])
    messages.success(request, "Módulo ativado com sucesso." if module.active else "Módulo desativado com sucesso.")
    return redirect(f"{reverse('core:system_screens')}?module={module.pk}")


@login_required
@role_required("TI")
def system_screen_edit(request, pk=None):
    screen = get_object_or_404(ScreenDefinition.objects.select_related("module"), pk=pk) if pk else None
    initial = {}
    if not screen:
        if request.GET.get("module", "").isdigit():
            initial["module"] = request.GET["module"]
        if request.GET.get("parent", "").isdigit():
            initial["parent"] = request.GET["parent"]
    module_id = (
        screen.module_id if screen
        else request.POST.get("module") or initial.get("module")
    )
    target_module = screen.module if screen else (
        Module.objects.filter(pk=module_id).first() if str(module_id or "").isdigit() else None
    )
    protected = bool(target_module and target_module.is_system)
    if protected and not screen:
        raise PermissionDenied("Não é permitido adicionar itens a um módulo estrutural.")
    if protected and request.method == "POST":
        raise PermissionDenied("Itens de módulos estruturais só podem ser alterados por migrações.")
    return_url = reverse("core:system_screens")
    if str(module_id or "").isdigit():
        return_url = f"{return_url}?{urlencode({'module': module_id})}"
    return_url = f"{return_url}#module-items"
    request.current_tab_title = "Global > Configuração do Sistema > Módulos e Telas > Configurar item"
    request.current_tab_root_title = "Módulos e Telas"
    request.current_module_title = "Global"
    request.current_return_url = return_url
    request.current_can_query = False
    request.current_can_remove = False
    request.current_new_url = ""
    request.current_can_save = not protected
    form = ScreenDefinitionForm(
        request.POST or None,
        instance=screen,
        initial=initial,
        protected=protected,
    )
    if request.method == "POST" and form.is_valid():
        saved_screen = form.save()
        messages.success(request, "Tela salva com sucesso.")
        return redirect("core:system_screen_edit", pk=saved_screen.pk)
    return render(request, "core/system_screen_form.html", {"form": form, "screen": screen})


def _sanitize_system_icon_svg(value):
    cleaned = bleach.clean(
        (value or "").strip(),
        tags={"svg", "g", "path", "circle", "ellipse", "rect", "line", "polyline", "polygon", "title", "desc"},
        attributes={
            "svg": {"xmlns", "viewBox", "viewbox", "preserveAspectRatio", "role", "aria-hidden", "focusable"},
            "*": {
                "class", "d", "x", "y", "x1", "x2", "y1", "y2", "cx", "cy", "r", "rx", "ry",
                "width", "height", "points", "fill", "fill-opacity", "fill-rule", "stroke", "stroke-opacity",
                "stroke-width", "stroke-linecap", "stroke-linejoin", "stroke-dasharray", "stroke-dashoffset",
                "stroke-miterlimit", "clip-rule", "opacity", "transform", "vector-effect",
            },
        },
        protocols=set(),
        strip=True,
    )
    return cleaned if cleaned.lstrip().lower().startswith("<svg") else ""


def _new_system_icon_code(name):
    base = _auxiliary_code(name).lower().replace("_", "-") or "icone"
    base = base[:50]
    candidate = base
    suffix = 2
    while IconeSistema.objects.filter(cd_icone=candidate).exists():
        marker = f"-{suffix}"
        candidate = f"{base[:50 - len(marker)]}{marker}"
        suffix += 1
    return candidate


@login_required
@role_required("TI")
def system_icons(request):
    request.current_tab_title = "Global > Configuração do Sistema > Módulos e Telas > Ícones"
    request.current_tab_root_title = "Ícones"
    request.current_module_title = "Global"
    request.current_can_query = True
    request.current_can_remove = True
    request.current_start_query = request.GET.get("consultar") != "1"
    if request.method == "POST":
        protected_icons = set(Module.objects.values_list("icon", flat=True)) | set(
            ScreenDefinition.objects.values_list("icon", flat=True)
        )
        with transaction.atomic():
            for icon in IconeSistema.objects.all():
                if request.POST.get(f"delete_{icon.pk}") == "1":
                    if icon.cd_icone in protected_icons:
                        icon.sn_ativo = False
                        icon.save(update_fields=("sn_ativo", "updated_at"))
                        messages.warning(
                            request,
                            f'O ícone "{icon.nm_icone}" está em uso e foi apenas inativado.',
                        )
                    else:
                        icon.delete()
                    continue
                if f"name_{icon.pk}" not in request.POST:
                    continue
                icon.nm_icone = request.POST.get(f"name_{icon.pk}", "").strip()
                icon.ds_svg = _sanitize_system_icon_svg(request.POST.get(f"svg_{icon.pk}", ""))
                icon.sn_ativo = request.POST.get(f"active_{icon.pk}") == "true"
                icon.save()
            new_names = request.POST.getlist("new_name")
            new_svgs = request.POST.getlist("new_svg")
            new_active = request.POST.getlist("new_active")
            for index, name in enumerate(new_names):
                name = name.strip()
                if not name:
                    continue
                IconeSistema.objects.create(
                    cd_icone=_new_system_icon_code(name),
                    nm_icone=name,
                    ds_svg=_sanitize_system_icon_svg(new_svgs[index] if index < len(new_svgs) else ""),
                    sn_ativo=index >= len(new_active) or new_active[index] == "true",
                )
        messages.success(request, "Ícones salvos com sucesso.")
        return redirect(f"{request.path}?consultar=1")
    icons = IconeSistema.objects.all()
    query = request.GET.get("q", "").strip().replace("%", "")
    if query:
        icons = icons.filter(
            Q(cd_icone__icontains=query) | Q(nm_icone__icontains=query) | Q(ds_svg__icontains=query)
        )
    icons = paginate_table(
        request,
        icons,
        {"cd_icone", "nm_icone", "sn_ativo"},
        "cd_icone",
    )
    return render(request, "core/system_icons.html", {"registros": icons})


@login_required
@role_required("TI")
@transaction.atomic
def system_navigation_reorder(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método não permitido."}, status=405)
    node_id = str(request.POST.get("node") or "").strip()
    parent_id = str(request.POST.get("parent") or "").strip()
    ordered_ids = [value for value in request.POST.getlist("order") if str(value).isdigit()]
    if not node_id.isdigit():
        return JsonResponse({"ok": False, "error": "Item inválido."}, status=400)
    node = get_object_or_404(ScreenDefinition.objects.select_related("module"), pk=node_id)
    if node.module.is_system:
        raise PermissionDenied("Itens de módulos estruturais não podem ser reordenados.")
    parent = None
    if parent_id:
        if not parent_id.isdigit():
            return JsonResponse({"ok": False, "error": "Grupo inválido."}, status=400)
        parent = get_object_or_404(
            ScreenDefinition,
            pk=parent_id,
            module=node.module,
            screen_type=ScreenDefinition.TYPE_GROUP,
        )
        if parent.pk == node.pk:
            return JsonResponse({"ok": False, "error": "Um item não pode conter a si mesmo."}, status=400)
    node.parent = parent
    node.parent_label = parent.title if parent else ""
    node.save(update_fields=["parent", "parent_label", "updated_at"])
    siblings = ScreenDefinition.objects.filter(module=node.module, parent=parent)
    valid_ids = set(siblings.values_list("pk", flat=True))
    normalized_ids = [int(value) for value in ordered_ids if int(value) in valid_ids]
    normalized_ids.extend(sibling_id for sibling_id in valid_ids if sibling_id not in normalized_ids)
    for position, sibling_id in enumerate(normalized_ids, start=1):
        ScreenDefinition.objects.filter(pk=sibling_id).update(order=position * 10)
    return JsonResponse({"ok": True})


@login_required
@role_required("TI")
def system_fields(request):
    fields = ScreenDefinition.objects.prefetch_related("fields").select_related("module").all()
    return render(request, "core/system_fields.html", {"screens": fields})


@login_required
@role_required("TI")
def system_field_edit(request, pk=None):
    field = get_object_or_404(ScreenField.objects.select_related("screen__module"), pk=pk) if pk else None
    if field and field.screen.module.is_system:
        raise PermissionDenied("Campos de módulos estruturais não podem ser alterados.")
    form = ScreenFieldForm(request.POST or None, instance=field)
    if request.method == "POST" and form.is_valid():
        saved_field = form.save()
        messages.success(request, "Campo salvo com sucesso.")
        return redirect("core:system_field_edit", pk=saved_field.pk)
    return render(request, "core/system_field_form.html", {"form": form, "field": field})


@login_required
def system_companies(request):
    request.current_can_query = True
    request.current_can_remove = True
    empresas = Empresa.objects.all()
    query = _query_text(request)
    if query:
        filtro = Q(nm_empresa__icontains=query) | Q(nr_cnpj__icontains=query) | Q(ds_cidade__icontains=query)
        if query.isdigit():
            filtro |= Q(cd_empresa=int(query))
        empresas = empresas.filter(filtro)
    empresas = paginate_table(
        request,
        empresas,
        {"cd_empresa", "nm_empresa", "nr_cnpj", "ds_cidade", "sg_estado", "sn_ativo"},
        "cd_empresa",
    )
    if request.method == "POST":
        for empresa in empresas:
            if request.POST.get(f"delete_{empresa.pk}") == "1":
                if empresa.cd_empresa != 1:
                    empresa.delete()
                continue
            if f"name_{empresa.pk}" not in request.POST:
                continue
            empresa.nm_empresa = request.POST.get(f"name_{empresa.pk}", empresa.nm_empresa)
            empresa.nr_cnpj = request.POST.get(f"cnpj_{empresa.pk}", empresa.nr_cnpj)
            empresa.ds_cidade = request.POST.get(f"city_{empresa.pk}", empresa.ds_cidade)
            empresa.sg_estado = request.POST.get(f"state_{empresa.pk}", empresa.sg_estado)
            empresa.sn_ativo = request.POST.get(f"active_{empresa.pk}") == "true"
            empresa.save()
        new_names = request.POST.getlist("new_name")
        next_code = (Empresa.objects.aggregate(max_code=Max("cd_empresa"))["max_code"] or 0) + 1
        for index, name in enumerate(new_names):
            name = name.strip()
            if not name:
                continue
            Empresa.objects.create(
                cd_empresa=next_code,
                nm_empresa=name,
                nr_cnpj=request.POST.getlist("new_cnpj")[index].strip() if index < len(request.POST.getlist("new_cnpj")) else "",
                ds_cidade=request.POST.getlist("new_city")[index].strip() if index < len(request.POST.getlist("new_city")) else "",
                sg_estado=request.POST.getlist("new_state")[index].strip() if index < len(request.POST.getlist("new_state")) else "",
                sn_ativo=True,
            )
            next_code += 1
        messages.success(request, "Empresas salvas com sucesso.")
        return redirect(f"{request.path}?consultar=1")
    return render(request, "core/system_companies.html", {"empresas": empresas})


@login_required
def system_company_edit(request, pk=None):
    empresa = get_object_or_404(Empresa, pk=pk) if pk else None
    form = EmpresaForm(request.POST or None, instance=empresa)
    if request.method == "POST" and form.is_valid():
        saved_empresa = form.save(commit=False)
        if not saved_empresa.pk:
            saved_empresa.cd_usuario_criacao = request.user
        saved_empresa.cd_usuario_atualizacao = request.user
        saved_empresa.save()
        messages.success(request, "Empresa salva com sucesso.")
        return redirect("core:system_company_edit", pk=saved_empresa.pk)
    return render(request, "core/system_company_form.html", {"form": form, "empresa": empresa})


@login_required
@role_required("TI")
def setores(request, tipo=None):
    empresa = get_object_or_404(Empresa, pk=request.session.get("cd_empresa") or 1)
    request.current_can_query = True
    request.current_can_remove = True
    request.current_tab_title = "Global > Empresa > Setores de Atendimento" if tipo == Setor.TipoSetor.ATENDIMENTO else "Global > Empresa > Setores"
    request.current_tab_root_title = "Setores de Atendimento" if tipo == Setor.TipoSetor.ATENDIMENTO else "Setores"
    request.current_module_title = "Global"
    registros = Setor.objects.filter(cd_empresa=empresa)
    if tipo:
        registros = registros.filter(tp_setor=tipo)
    query = _query_text(request)
    if query:
        registros = registros.filter(Q(nm_setor__icontains=query) | Q(ds_observacao__icontains=query))
    registros = paginate_table(
        request,
        registros,
        {"cd_setor", "nm_setor", "tp_setor", "sn_ativo"},
        "cd_setor",
    )
    if request.method == "POST":
        for setor in registros:
            if request.POST.get(f"delete_{setor.pk}") == "1":
                setor.sn_ativo = False
                setor.save(update_fields=["sn_ativo", "dh_atualizacao"])
                continue
            if f"name_{setor.pk}" not in request.POST:
                continue
            setor.nm_setor = request.POST.get(f"name_{setor.pk}", setor.nm_setor).strip()
            setor.tp_setor = tipo or request.POST.get(f"type_{setor.pk}", setor.tp_setor)
            setor.ds_observacao = request.POST.get(f"note_{setor.pk}", setor.ds_observacao).strip()
            setor.sn_ativo = request.POST.get(f"active_{setor.pk}") == "true"
            setor.save()
        new_names = request.POST.getlist("new_name")
        new_types = request.POST.getlist("new_type")
        new_notes = request.POST.getlist("new_note")
        for index, name in enumerate(new_names):
            name = name.strip()
            if not name:
                continue
            Setor.objects.update_or_create(
                cd_empresa=empresa,
                tp_setor=tipo or (new_types[index] if index < len(new_types) and new_types[index] else Setor.TipoSetor.EMPRESA),
                nm_setor=name,
                defaults={
                    "ds_observacao": new_notes[index].strip() if index < len(new_notes) else "",
                    "sn_ativo": True,
                },
            )
        messages.success(request, "Setores salvos com sucesso.")
        return redirect(f"{request.path}?consultar=1")
    return render(request, "core/setores.html", {"registros": registros, "tipo": tipo, "tipos": Setor.TipoSetor.choices})


@login_required
def global_auxiliary_values(request, tabela, custom_title=None, custom_module=None):
    labels = {
        "tipo_sanguineo": "Tipo sanguíneo",
        "sexo": "Sexo",
        "estado_civil": "Estado civil",
        "naturalidade": "Naturalidade",
        "nacionalidade": "Nacionalidade",
        "cidade": "Cidade",
        "estado": "Estado",
        "cep": "CEPs",
        "bairro": "Bairros",
        "tipo_logradouro": "Tipos de Logradouro",
        "especialidade": "Especialidades",
        "conselho_profissional": "Conselhos Profissionais",
        "orgao_emissor": "Órgãos Emissores",
        "banco": "Bancos",
        "pais": "Nacionalidades",
        "tipo_prestador": "Tipos de Prestador",
        "tipo_vinculo": "Tipos de Vínculo",
        "motivo_alteracao": "Motivos de alteração",
        "cids": "CIDs",
        "motivos_alta": "Motivos de alta",
    }
    label = custom_title or labels.get(tabela, tabela.replace("_", " ").title())
    request.current_tab_title = (
        f"{custom_module} > {label}"
        if custom_module
        else f"Global > Tabelas > Auxiliares > {label}"
    )
    request.current_tab_root_title = label
    request.current_module_title = custom_module or "Global"
    request.current_can_query = True
    request.current_can_remove = True
    tabela_auxiliar, _ = TabelaAuxiliarGlobal.objects.get_or_create(
        ds_tabela=tabela,
        defaults={"ds_descricao": label, "sn_ativo": True},
    )
    query = _query_text(request)
    if request.method == "POST":
        for valor in tabela_auxiliar.valores.all():
            if request.POST.get(f"delete_{valor.pk}") == "1":
                try:
                    valor.delete()
                except ProtectedError:
                    messages.error(
                        request,
                        f"{valor.ds_valor} está em uso e não pode ser excluído.",
                    )
                continue
            if f"description_{valor.pk}" not in request.POST:
                continue
            valor.ds_valor = request.POST.get(f"description_{valor.pk}", valor.ds_valor)
            valor.ds_grupo = request.POST.get(f"group_{valor.pk}", valor.ds_grupo)
            valor.sn_ativo = request.POST.get(f"active_{valor.pk}") == "true"
            valor.save()
        new_descriptions = request.POST.getlist("new_description")
        new_groups = request.POST.getlist("new_group")
        new_actives = request.POST.getlist("new_active")
        created = 0
        for index, new_description in enumerate(new_descriptions):
            new_description = new_description.strip()
            if not new_description:
                continue
            new_code = _auxiliary_code(new_description)
            ValorAuxiliarGlobal.objects.update_or_create(
                cd_tabela_auxiliar_global=tabela_auxiliar,
                cd_valor=new_code,
                defaults={
                    "ds_valor": new_description,
                    "ds_grupo": new_groups[index].strip() if index < len(new_groups) else "",
                    "sn_ativo": (new_actives[index] if index < len(new_actives) else "true") == "true",
                },
            )
            created += 1
        if new_descriptions and not created and not tabela_auxiliar.valores.exists():
            messages.error(request, "Informe a descrição obrigatória antes de salvar.")
        else:
            messages.success(request, "Tabela auxiliar salva com sucesso.")
        return redirect(f"{request.path}?consultar=1")
    valores = tabela_auxiliar.valores.all()
    if query:
        value_filter = Q(cd_valor__icontains=query) | Q(ds_valor__icontains=query) | Q(ds_grupo__icontains=query)
        if query.isdigit():
            value_filter |= Q(cd_valor_auxiliar_global=int(query))
        valores = valores.filter(value_filter)
    valores = paginate_table(
        request,
        valores,
        {"cd_valor_auxiliar_global", "cd_valor", "ds_valor", "ds_grupo", "sn_ativo"},
        "cd_valor_auxiliar_global",
    )
    estados = []
    if tabela == "cidade":
        estados = list(
            ValorAuxiliarGlobal.objects.filter(
                cd_tabela_auxiliar_global__ds_tabela="estado",
                sn_ativo=True,
            ).order_by("ds_valor")
        )
    return render(
        request,
        "core/global_auxiliary_values.html",
        {"tabela": tabela_auxiliar, "valores": valores, "estados": estados},
    )


@login_required
@role_required("TI")
def global_tables(request):
    request.current_tab_title = "Global > Tabelas Auxiliares"
    request.current_tab_root_title = "Tabelas Auxiliares"
    request.current_module_title = "Global"
    tables = TabelaAuxiliarGlobal.objects.filter(sn_ativo=True).order_by("ds_descricao", "ds_tabela")
    return render(request, "core/global_tables.html", {"tables": tables})


@login_required
@role_required("TI")
def global_ceps(request):
    request.current_tab_title = "Global > CEPs"
    request.current_tab_root_title = "CEPs"
    request.current_module_title = "Global"
    request.current_can_query = True
    request.current_can_remove = False
    records = Cep.objects.all()
    query = _query_text(request)
    if query:
        digits = "".join(character for character in query if character.isdigit())
        cep_filter = (
            Q(ds_logradouro__icontains=query)
            | Q(ds_bairro__icontains=query)
            | Q(ds_cidade__icontains=query)
            | Q(cd_cidade__icontains=query)
            | Q(tp_logradouro__icontains=query)
            | Q(sg_estado__icontains=query)
        )
        if digits:
            cep_filter |= Q(nr_cep__icontains=digits)
        records = records.filter(cep_filter)
    if request.method == "POST":
        for record in Cep.objects.all():
            if request.POST.get(f"delete_{record.pk}") == "1":
                try:
                    record.delete()
                except ProtectedError:
                    record.sn_ativo = False
                    record.save(update_fields=["sn_ativo", "updated_at"])
                continue
            if f"nr_cep_{record.pk}" not in request.POST and f"postal_code_{record.pk}" not in request.POST:
                continue
            record.nr_cep = "".join(
                character
                for character in request.POST.get(f"nr_cep_{record.pk}", request.POST.get(f"postal_code_{record.pk}", ""))
                if character.isdigit()
            )[:8]
            record.sg_estado = request.POST.get(f"sg_estado_{record.pk}", request.POST.get(f"state_{record.pk}", "")).strip().upper()[:2]
            record.cd_cidade = request.POST.get(f"cd_cidade_{record.pk}", request.POST.get(f"city_code_{record.pk}", "")).strip()[:40]
            record.ds_cidade = request.POST.get(f"ds_cidade_{record.pk}", request.POST.get(f"city_{record.pk}", "")).strip()[:160]
            record.tp_logradouro = request.POST.get(f"tp_logradouro_{record.pk}", request.POST.get(f"street_type_{record.pk}", "")).strip()[:40]
            record.ds_logradouro = request.POST.get(f"ds_logradouro_{record.pk}", request.POST.get(f"street_{record.pk}", "")).strip()[:220]
            record.ds_bairro = request.POST.get(f"ds_bairro_{record.pk}", request.POST.get(f"district_{record.pk}", "")).strip()[:160]
            record.sn_ativo = request.POST.get(f"sn_ativo_{record.pk}", request.POST.get(f"active_{record.pk}")) == "true"
            record.save()
        new_codes = request.POST.getlist("new_nr_cep") or request.POST.getlist("new_postal_code")
        for index, postal_code in enumerate(new_codes):
            digits = "".join(character for character in postal_code if character.isdigit())[:8]
            if not digits:
                continue
            new_states = request.POST.getlist("new_sg_estado") or request.POST.getlist("new_state")
            new_city_codes = request.POST.getlist("new_cd_cidade") or request.POST.getlist("new_city_code")
            new_cities = request.POST.getlist("new_ds_cidade") or request.POST.getlist("new_city")
            new_street_types = request.POST.getlist("new_tp_logradouro") or request.POST.getlist("new_street_type")
            new_streets = request.POST.getlist("new_ds_logradouro") or request.POST.getlist("new_street")
            new_districts = request.POST.getlist("new_ds_bairro") or request.POST.getlist("new_district")
            new_actives = request.POST.getlist("new_sn_ativo") or request.POST.getlist("new_active")
            Cep.objects.update_or_create(
                nr_cep=digits,
                defaults={
                    "sg_estado": (new_states[index] if index < len(new_states) else "").strip().upper()[:2],
                    "cd_cidade": (new_city_codes[index] if index < len(new_city_codes) else "").strip()[:40],
                    "ds_cidade": (new_cities[index] if index < len(new_cities) else "").strip()[:160],
                    "tp_logradouro": (new_street_types[index] if index < len(new_street_types) else "").strip()[:40],
                    "ds_logradouro": (new_streets[index] if index < len(new_streets) else "").strip()[:220],
                    "ds_bairro": (new_districts[index] if index < len(new_districts) else "").strip()[:160],
                    "sn_ativo": (new_actives[index] if index < len(new_actives) else "true") == "true",
                },
            )
        messages.success(request, "CEPs salvos com sucesso.")
        return redirect(f"{request.path}?consultar=1")
    records = paginate_table(
        request,
        records,
        {"cd_cep", "nr_cep", "sg_estado", "ds_cidade", "ds_logradouro", "ds_bairro", "sn_ativo"},
        "cd_cep",
    )
    auxiliary_values = ValorAuxiliarGlobal.objects.filter(sn_ativo=True).select_related("cd_tabela_auxiliar_global")
    estados = auxiliary_values.filter(cd_tabela_auxiliar_global__ds_tabela="estado").order_by("ds_valor")
    cidades = auxiliary_values.filter(cd_tabela_auxiliar_global__ds_tabela="cidade").order_by("ds_valor")
    tipos_logradouro = auxiliary_values.filter(cd_tabela_auxiliar_global__ds_tabela="tipo_logradouro").order_by("ds_valor")
    return render(
        request,
        "core/global_ceps.html",
        {
            "records": records,
            "estados": estados,
            "cidades": cidades,
            "tipos_logradouro": tipos_logradouro,
        },
    )


IMPORT_TABLES = {
    "cep": "CEPs",
    "estado": "Estados",
    "cidade": "Cidades",
    "tipo_logradouro": "Tipos de Logradouro",
}


def _import_rows(upload):
    suffix = upload.name.rsplit(".", 1)[-1].lower()
    content = upload.read()
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=";,")
        except csv.Error:
            dialect = csv.excel_semicolon
        return list(csv.DictReader(StringIO(text), dialect=dialect))
    if suffix == "xlsx":
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as error:
            raise ValueError("A importação XLSX requer a dependência openpyxl instalada.") from error
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, [])]
        return [dict(zip(headers, values)) for values in rows]
    raise ValueError("Formato não suportado. Utilize CSV ou XLSX.")


def _normalized_import_row(row):
    normalized = {}
    for key, value in row.items():
        normalized_key = _auxiliary_code(str(key or "")).lower()
        normalized[normalized_key] = str(value or "").strip()
    return normalized


@login_required
@role_required("TI")
def global_integrations(request):
    request.current_tab_title = "Global > Integrações"
    request.current_tab_root_title = "Importação de dados"
    request.current_module_title = "Global"
    report = None
    if request.method == "POST":
        table_name = request.POST.get("table_name", "")
        upload = request.FILES.get("file")
        errors = []
        processed = 0
        created = 0
        updated = 0
        if table_name not in IMPORT_TABLES:
            errors.append("Selecione uma integração válida.")
        if not upload:
            errors.append("Selecione um arquivo CSV ou XLSX.")
        if not errors:
            try:
                rows = _import_rows(upload)
                with transaction.atomic():
                    for row_number, source_row in enumerate(rows, start=2):
                        row = _normalized_import_row(source_row)
                        description = row.get("descricao") or row.get("nome") or row.get("logradouro")
                        code = row.get("codigo") or row.get("cep") or row.get("sigla")
                        group = row.get("grupo", "")
                        if table_name == "cidade":
                            group = row.get("uf") or row.get("estado") or group
                        elif table_name == "bairro":
                            group = row.get("cidade") or group
                        elif table_name == "cep":
                            state = row.get("uf") or row.get("estado")
                            city = row.get("cidade")
                            group = "|".join(part for part in (state, city) if part) or group
                        if not description:
                            errors.append(f"Linha {row_number}: descrição obrigatória.")
                            continue
                        code = (code or _auxiliary_code(description))[:40]
                        if not code:
                            errors.append(f"Linha {row_number}: código inválido.")
                            continue
                        active = row.get("ativo", "SIM").upper() not in {"0", "NAO", "NÃO", "FALSE", "INATIVO"}
                        if table_name == "cep":
                            digits = "".join(character for character in (row.get("cep") or code) if character.isdigit())[:8]
                            if not digits:
                                errors.append(f"Linha {row_number}: CEP inválido.")
                                continue
                            _, was_created = Cep.objects.update_or_create(
                                nr_cep=digits,
                                defaults={
                                    "sg_estado": (row.get("uf") or row.get("estado") or "")[:2].upper(),
                                    "cd_cidade": (row.get("codigo_cidade") or row.get("cidade") or "")[:40],
                                    "ds_cidade": (row.get("cidade") or "")[:160],
                                    "tp_logradouro": (row.get("tipo_logradouro") or "")[:40],
                                    "ds_logradouro": description[:220],
                                    "ds_bairro": (row.get("bairro") or "")[:160],
                                    "sn_ativo": active,
                                },
                            )
                        else:
                            table, _ = TabelaAuxiliarGlobal.objects.get_or_create(
                                ds_tabela=table_name,
                                defaults={"ds_descricao": IMPORT_TABLES[table_name], "sn_ativo": True},
                            )
                            _, was_created = ValorAuxiliarGlobal.objects.update_or_create(
                                cd_tabela_auxiliar_global=table,
                                cd_valor=code,
                                defaults={
                                    "ds_valor": description[:160],
                                    "ds_grupo": group[:40],
                                    "sn_ativo": active,
                                },
                            )
                        processed += 1
                        created += int(was_created)
                        updated += int(not was_created)
            except (ValueError, UnicodeDecodeError, StopIteration) as error:
                errors.append(str(error))
            except Exception as error:
                errors.append(f"Falha ao processar o arquivo: {error}")
        report = {
            "processed": processed,
            "created": created,
            "updated": updated,
            "errors": errors,
        }
        if processed:
            messages.success(request, f"Importação concluída: {processed} registro(s) processado(s).")
        elif errors:
            messages.error(request, "A importação não processou registros.")
    return render(
        request,
        "core/global_integrations.html",
        {"import_tables": IMPORT_TABLES, "report": report},
    )




@login_required
def tipo_prestador_conselho(request):
    request.current_tab_title = "Global > Tabelas > Tipo de prestador x conselho"
    request.current_tab_root_title = "Tipo de prestador x conselho"
    request.current_module_title = "Global"
    request.current_can_query = True
    request.current_can_remove = True
    registros = TipoPrestadorConselho.objects.all()
    query = _query_text(request)
    if query:
        registros = registros.filter(Q(tp_prestador__icontains=query) | Q(ds_conselho__icontains=query))
    registros = paginate_table(
        request,
        registros,
        {"id", "tp_prestador", "ds_conselho", "sn_ativo"},
        "id",
    )
    tipos = ValorAuxiliarGlobal.objects.filter(
        cd_tabela_auxiliar_global__ds_tabela="tipo_prestador",
        sn_ativo=True,
    ).order_by("ds_valor")
    if request.method == "POST":
        for registro in registros:
            if request.POST.get(f"delete_{registro.pk}") == "1":
                registro.sn_ativo = False
                registro.save(update_fields=["sn_ativo", "updated_at"])
                continue
            if f"type_{registro.pk}" not in request.POST:
                continue
            registro.tp_prestador = request.POST.get(f"type_{registro.pk}", registro.tp_prestador)
            registro.ds_conselho = request.POST.get(f"council_{registro.pk}", registro.ds_conselho).strip().upper()
            registro.sn_ativo = request.POST.get(f"active_{registro.pk}") == "true"
            registro.save()
        councils = request.POST.getlist("new_council")
        for index, provider_type in enumerate(request.POST.getlist("new_type")):
            provider_type = provider_type.strip()
            if not provider_type or index >= len(councils) or not councils[index].strip():
                continue
            TipoPrestadorConselho.objects.update_or_create(
                tp_prestador=provider_type,
                defaults={"ds_conselho": councils[index].strip().upper(), "sn_ativo": True},
            )
        messages.success(request, "Vínculos de prestador e conselho salvos com sucesso.")
        return redirect(f"{request.path}?consultar=1")
    return render(request, "core/tipo_prestador_conselho.html", {"registros": registros, "tipos": tipos})


@login_required
def city_options(request):
    uf = request.GET.get("uf", "")
    cidades = ValorAuxiliarGlobal.objects.filter(
        cd_tabela_auxiliar_global__ds_tabela="cidade",
        sn_ativo=True,
        ds_grupo=uf,
    ).order_by("ds_valor")
    return JsonResponse({"cidades": [{"value": cidade.cd_valor, "label": cidade.ds_valor} for cidade in cidades]})


@login_required
def cep_option(request):
    raw_value = request.GET.get("cep", "")
    value = Cep.objects.filter(sn_ativo=True)
    value = value.filter(pk=int(raw_value)) if raw_value.isdigit() and len(raw_value) < 8 else value.filter(
        nr_cep="".join(character for character in raw_value if character.isdigit())
    )
    value = value.first()
    if not value:
        return JsonResponse({"estado": "", "cidade": "", "bairro": "", "logradouro": "", "tipo_logradouro": ""})
    return JsonResponse(
        {
            "estado": value.sg_estado,
            "cidade": value.cd_cidade or value.ds_cidade,
            "bairro": value.ds_bairro,
            "logradouro": value.ds_logradouro,
            "tipo_logradouro": value.tp_logradouro,
        }
    )
